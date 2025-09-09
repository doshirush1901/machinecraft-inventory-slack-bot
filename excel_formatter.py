#!/usr/bin/env python3
"""
Excel Formatter for Professional Inventory Database
Adds McMaster-Carr style formatting and formulas
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import warnings
warnings.filterwarnings('ignore')

class ExcelFormatter:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        
    def format_worksheet(self, sheet_name):
        """Format a worksheet with professional styling"""
        if sheet_name not in self.workbook.sheetnames:
            return
            
        ws = self.workbook[sheet_name]
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        if ws.max_row > 0:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format data rows
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                
                # Right align numeric columns
                if col > 4:  # Price and quantity columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Auto-adjust column widths
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, ws.max_row + 1):
                cell_value = str(ws.cell(row=row, column=col).value or '')
                max_length = max(max_length, len(cell_value))
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = 'A2'
        
        # Add filters
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    
    def add_conditional_formatting(self, sheet_name):
        """Add conditional formatting for better visualization"""
        if sheet_name not in self.workbook.sheetnames:
            return
            
        ws = self.workbook[sheet_name]
        
        # Find price column
        price_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value and 'price' in str(ws.cell(row=1, column=col).value).lower():
                price_col = col
                break
        
        if price_col and ws.max_row > 1:
            # Color scale for prices
            price_range = f"{get_column_letter(price_col)}2:{get_column_letter(price_col)}{ws.max_row}"
            color_scale = ColorScaleRule(
                start_type='min', start_color='FF6B6B',
                mid_type='percentile', mid_value=50, mid_color='FFE66D',
                end_type='max', end_color='4ECDC4'
            )
            ws.conditional_formatting.add(price_range, color_scale)
        
        # Find quantity column
        qty_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value and 'quantity' in str(ws.cell(row=1, column=col).value).lower():
                qty_col = col
                break
        
        if qty_col and ws.max_row > 1:
            # Data bars for quantities
            qty_range = f"{get_column_letter(qty_col)}2:{get_column_letter(qty_col)}{ws.max_row}"
            data_bar = DataBarRule(
                start_type='min', start_value=0,
                end_type='max', end_value=None,
                color='4F81BD'
            )
            ws.conditional_formatting.add(qty_range, data_bar)
    
    def add_summary_formulas(self, sheet_name):
        """Add summary formulas to the worksheet"""
        if sheet_name not in self.workbook.sheetnames:
            return
            
        ws = self.workbook[sheet_name]
        
        # Add summary row at the bottom
        summary_row = ws.max_row + 2
        
        # Find key columns
        price_col = None
        qty_col = None
        total_col = None
        
        for col in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row=1, column=col).value or '').lower()
            if 'price' in cell_value and 'unit' in cell_value:
                price_col = col
            elif 'quantity' in cell_value:
                qty_col = col
            elif 'total' in cell_value and 'value' in cell_value:
                total_col = col
        
        # Add summary formulas
        if price_col:
            ws.cell(row=summary_row, column=price_col, value="Average Price:")
            ws.cell(row=summary_row, column=price_col + 1, value=f"=AVERAGE({get_column_letter(price_col)}2:{get_column_letter(price_col)}{ws.max_row-1})")
            
        if qty_col:
            ws.cell(row=summary_row + 1, column=qty_col, value="Total Quantity:")
            ws.cell(row=summary_row + 1, column=qty_col + 1, value=f"=SUM({get_column_letter(qty_col)}2:{get_column_letter(qty_col)}{ws.max_row-1})")
            
        if total_col:
            ws.cell(row=summary_row + 2, column=total_col, value="Total Value:")
            ws.cell(row=summary_row + 2, column=total_col + 1, value=f"=SUM({get_column_letter(total_col)}2:{get_column_letter(total_col)}{ws.max_row-1})")
    
    def format_all_sheets(self):
        """Format all sheets in the workbook"""
        for sheet_name in self.workbook.sheetnames:
            print(f"Formatting sheet: {sheet_name}")
            self.format_worksheet(sheet_name)
            self.add_conditional_formatting(sheet_name)
            
            # Add summary formulas for main sheets
            if sheet_name in ['Master Inventory', 'Category Analysis', 'Brand Analysis']:
                self.add_summary_formulas(sheet_name)
    
    def save_formatted_file(self, output_path):
        """Save the formatted workbook"""
        self.workbook.save(output_path)
        print(f"Formatted file saved as: {output_path}")

def main():
    input_file = "/Users/rushabhdoshi/Library/CloudStorage/Box-Box/MCRAFT 2023/11 Inventory/Professional_Master_Inventory.xlsx"
    output_file = "/Users/rushabhdoshi/Library/CloudStorage/Box-Box/MCRAFT 2023/11 Inventory/Machinecraft_Professional_Inventory_Database.xlsx"
    
    print("Starting Excel formatting...")
    formatter = ExcelFormatter(input_file)
    formatter.format_all_sheets()
    formatter.save_formatted_file(output_file)
    print("Excel formatting complete!")

if __name__ == "__main__":
    main()
